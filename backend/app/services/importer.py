import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, insert

from ..models import StagingRow
from ..sheet_schemas import (
    DATA_COLUMNS,
    DATA_SHEETS,
    PRODUCT_FIELDS,
    PRODUCT_SHEETS,
    REPEATED_INVENTORY_HEADERS,
)

FILENAME_DATE_RE = re.compile(
    r"(\d{4})[.\-_/年]?(\d{1,2})[.\-_/月]?(\d{1,2})"
)
DATE_HEADER_RE = re.compile(r"^\d{1,2}[.\-/月]\d{1,2}(日)?(销量)?$")


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").replace("\n", "").strip()
    return text


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    text = str(value).replace(",", "").strip()
    if text in ("", "-", "--", "#N/A", "#VALUE!", "无动销"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_scalar(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    return text if text else None


def detect_date_from_filename(filename: str):
    match = FILENAME_DATE_RE.search(filename)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    return date.today().isoformat()


def find_column(headers, aliases):
    normalized_aliases = {normalize_header(a) for a in aliases}
    for idx, header in enumerate(headers):
        if normalize_header(header) in normalized_aliases:
            return idx
    return None


def find_dynamic_date_column(headers):
    for idx, header in enumerate(headers):
        if DATE_HEADER_RE.match(normalize_header(header)):
            return idx
    return None


def build_row_dict(headers, row):
    result = {}
    for idx, value in enumerate(row):
        header = normalize_header(headers[idx]) if idx < len(headers) else f"col{idx + 1}"
        parsed = parse_scalar(value)
        if parsed is None or parsed == "":
            continue
        if header in result:
            header = f"{header}_{idx + 1}"
        result[header] = parsed
    return result


def open_workbook(path):
    return load_workbook(path, read_only=True, data_only=True)


def open_formula_workbook(path):
    return load_workbook(path, read_only=True, data_only=False)


def _data_sheet_rows(ws, sheet_name, snapshot_date):
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    spec = DATA_COLUMNS[sheet_name]
    key_idx = find_column(headers, spec["key"])
    if key_idx is None:
        return []

    inventory_idx = find_column(headers, spec.get("inventory", []))
    yesterday_idx = find_column(headers, spec.get("yesterday", []))
    if yesterday_idx is None and sheet_name in ("浪莎童装自营", "浪莎童装唯品"):
        yesterday_idx = find_dynamic_date_column(headers)
    seven_idx = find_column(headers, spec.get("seven", []))
    thirty_idx = find_column(headers, spec.get("thirty", []))

    rows = []
    for row_number, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(value is not None for value in row):
            continue
        key_value = row[key_idx]
        if key_value is None:
            continue
        sku = str(key_value).strip()
        if not sku:
            continue
        rows.append(
            {
                "sheet_name": sheet_name,
                "row_number": row_number,
                "kind": "data",
                "date": snapshot_date,
                "sku": sku,
                "inventory": parse_number(row[inventory_idx]) if inventory_idx is not None else None,
                "yesterday_sales": parse_number(row[yesterday_idx]) if yesterday_idx is not None else None,
                "seven_sales": parse_number(row[seven_idx]) if seven_idx is not None else None,
                "thirty_sales": parse_number(row[thirty_idx]) if thirty_idx is not None else None,
                "raw_json": json.dumps(build_row_dict(headers, row), ensure_ascii=False),
            }
        )
    return rows


def _find_header_row(ws):
    required = {"条形码", "产品编号", "产品名称"}
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1
    ):
        headers = [normalize_header(v) for v in row]
        if required & set(headers):
            return row_number
    return 1


def _inventory_runs(headers):
    runs = []
    current = []
    for idx, header in enumerate(headers):
        if normalize_header(header) in REPEATED_INVENTORY_HEADERS:
            current.append(idx)
        else:
            if len(current) >= 2:
                runs.append(current)
            current = []
    if len(current) >= 2:
        runs.append(current)
    return runs


def _product_sheet_rows(ws, sheet_name, snapshot_date, formula_ws=None):
    header_row_number = _find_header_row(ws)
    headers = next(
        ws.iter_rows(min_row=header_row_number, max_row=header_row_number, values_only=True)
    )
    field_idx = {}
    for field, aliases in PRODUCT_FIELDS.items():
        idx = find_column(headers, aliases)
        if idx is not None:
            field_idx[field] = idx
    brand_idx = find_column(headers, ["品牌"])

    rows = []
    inventory_runs = _inventory_runs(headers)
    formula_columns = set()
    if formula_ws is not None:
        for row_number, row in enumerate(
            formula_ws.iter_rows(
                min_row=header_row_number + 1,
                max_row=header_row_number + 8,
                values_only=True,
            ),
            start=header_row_number + 1,
        ):
            if not any(value is not None for value in row):
                continue
            for idx, value in enumerate(row):
                if isinstance(value, str) and value.startswith("="):
                    formula_columns.add(idx)
    if formula_columns:
        inventory_runs = [
            [idx for idx in run if idx not in formula_columns]
            for run in inventory_runs
            if [idx for idx in run if idx not in formula_columns]
        ]
    for row_number, row in enumerate(
        ws.iter_rows(
            min_row=header_row_number + 1, values_only=True
        ),
        start=header_row_number + 1,
    ):
        if not any(value is not None for value in row):
            continue
        barcode_value = row[field_idx["barcode"]] if "barcode" in field_idx else None
        if barcode_value is None:
            continue
        barcode = str(barcode_value).strip()
        if not barcode:
            continue

        raw = build_row_dict(headers, row)
        product = {
            "sheet_name": sheet_name,
            "row_number": row_number,
            "kind": "master",
            "date": snapshot_date,
            "sku": barcode,
            "inventory": None,
            "yesterday_sales": None,
            "seven_sales": None,
            "thirty_sales": None,
            "raw_json": json.dumps(raw, ensure_ascii=False),
            "product_fields": {},
        }
        for field, idx in field_idx.items():
            value = row[idx]
            if field in ("sale_price", "purchase_price"):
                product["product_fields"][field] = parse_number(value)
            else:
                parsed = parse_scalar(value)
                product["product_fields"][field] = parsed
        if brand_idx is not None:
            product["product_fields"]["brand"] = parse_scalar(row[brand_idx])
        product["product_fields"].setdefault("brand", sheet_name.split("-")[0])
        rows.append(product)

        for run in inventory_runs:
            run_date = date.fromisoformat(snapshot_date)
            for offset, col_idx in enumerate(run):
                value = parse_number(row[col_idx])
                if value is None:
                    continue
                inventory_date = run_date - timedelta(days=len(run) - 1 - offset)
                rows.append(
                    {
                        "sheet_name": sheet_name,
                        "row_number": row_number,
                        "kind": "inventory_history",
                        "date": inventory_date.isoformat(),
                        "sku": barcode,
                        "inventory": value,
                        "yesterday_sales": None,
                        "seven_sales": None,
                        "thirty_sales": None,
                        "raw_json": json.dumps({"header": normalize_header(headers[col_idx])}, ensure_ascii=False),
                        "product_fields": None,
                    }
                )
    return rows


def parse_workbook(path, snapshot_date=None, include_master=False):
    path = Path(path)
    if snapshot_date is None:
        snapshot_date = detect_date_from_filename(path.name)
    wb = open_workbook(path)
    formula_wb = open_formula_workbook(path) if include_master else None
    formula_sheets = {ws.title.strip(): ws for ws in formula_wb.worksheets} if formula_wb else {}
    sheet_summaries = []
    parsed_rows = []
    total_rows = 0

    for ws in wb.worksheets:
        name = ws.title.strip()
        is_data = name in DATA_SHEETS
        is_master = include_master and name in PRODUCT_SHEETS
        if not is_data and not is_master:
            continue
        if is_master:
            rows = _product_sheet_rows(
                ws, name, snapshot_date, formula_ws=formula_sheets.get(name)
            )
            kind = "master"
        else:
            rows = _data_sheet_rows(ws, name, snapshot_date)
            kind = "data"
        total_rows += len(rows)
        sheet_summaries.append(
            {
                "sheet_name": name,
                "kind": kind,
                "row_count": len(rows),
            }
        )
        parsed_rows.extend(rows)
    wb.close()
    if formula_wb is not None:
        formula_wb.close()
    return snapshot_date, sheet_summaries, parsed_rows


def store_staging(db, token, rows):
    master_rows = [r for r in rows if r["kind"] == "master"]
    data_rows = [r for r in rows if r["kind"] != "master"]
    values = []
    for row in data_rows:
        values.append(
            {
                "token": token,
                "sheet_name": row["sheet_name"],
                "row_number": row["row_number"],
                "kind": row["kind"],
                "date": row["date"],
                "sku": row["sku"],
                "inventory": row["inventory"],
                "yesterday_sales": row["yesterday_sales"],
                "seven_sales": row["seven_sales"],
                "thirty_sales": row["thirty_sales"],
                "raw_json": row["raw_json"],
            }
        )
    for start in range(0, len(values), 20_000):
        db.execute(insert(StagingRow), values[start : start + 20_000])
        db.commit()

    product_values = []
    for row in master_rows:
        fields = row["product_fields"] or {}
        product_values.append(
            {
                "token": token,
                "sheet_name": row["sheet_name"],
                "row_number": row["row_number"],
                "kind": "master",
                "date": row["date"],
                "sku": row["sku"],
                "inventory": None,
                "yesterday_sales": None,
                "seven_sales": None,
                "thirty_sales": None,
                "raw_json": row["raw_json"],
                "product_fields_json": json.dumps(fields, ensure_ascii=False),
            }
        )
    if product_values:
        for start in range(0, len(product_values), 5_000):
            db.execute(insert(StagingRow), product_values[start : start + 5_000])
            db.commit()


def discard_staging(db, token):
    db.execute(delete(StagingRow).where(StagingRow.token == token))
    db.commit()
